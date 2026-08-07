import csv

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from apps.authentication.mixins import admin_required
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from apps.transaction.models import Peminjaman
from apps.master_data.models import Dosen, Laboratorium, Mahasiswa

EXPORT_HEADERS = [
    'No', 'Tanggal Pinjam', 'NIM', 'Nama Mahasiswa',
    'Program Studi', 'Dosen', 'Ruangan', 'Nomor Kunci',
    'Jam Pinjam', 'Tanggal Kembali', 'Jam Kembali', 'Keperluan', 'Status',
]

COLUMN_WIDTHS = {
    'A': 5, 'B': 15, 'C': 15, 'D': 25, 'E': 20, 'F': 25, 'G': 20,
    'H': 15, 'I': 12, 'J': 15, 'K': 12, 'L': 30, 'M': 15,
}


def _get_filtered_data(request):
    """Return filtered queryset dan parameter filter aktif."""
    qs = Peminjaman.objects.select_related(
        'mahasiswa', 'kunci', 'laboratorium', 'dosen'
    )

    tgl_awal = request.GET.get('tgl_awal')
    tgl_akhir = request.GET.get('tgl_akhir')
    status = request.GET.get('status')
    ruangan = request.GET.get('ruangan')
    dosen = request.GET.get('dosen')
    prodi = request.GET.get('prodi')

    if tgl_awal:
        qs = qs.filter(tanggal_pinjam__gte=tgl_awal)
    if tgl_akhir:
        qs = qs.filter(tanggal_pinjam__lte=tgl_akhir)
    if status:
        qs = qs.filter(status=status)
    if ruangan:
        qs = qs.filter(laboratorium_id=ruangan)
    if dosen:
        qs = qs.filter(dosen_id=dosen)
    if prodi:
        qs = qs.filter(mahasiswa__program_studi__icontains=prodi)

    return qs, tgl_awal, tgl_akhir, status, ruangan, dosen, prodi


def _fmt_date(d):
    return d.strftime('%d/%m/%Y') if d else '-'


def _fmt_time(t):
    return str(t)[:5] if t else '-'

def _prodi_list():
    return list(
        Mahasiswa.objects.exclude(program_studi='')
        .values_list('program_studi', flat=True)
        .distinct().order_by('program_studi')
    )


def _peminjaman_to_row(i, p):
    """Konversi satu record Peminjaman jadi list untuk export."""
    mhs = p.mahasiswa
    lab = p.laboratorium
    return [
        i,
        _fmt_date(p.tanggal_pinjam),
        mhs.nim if mhs else '-',
        mhs.nama if mhs else '-',
        mhs.program_studi if mhs else '-',
        p.dosen.nama if p.dosen else '-',
        f'{lab.kode_lab} - {lab.nama_lab}' if lab else '-',
        p.kunci.nomor_kunci if p.kunci else '-',
        _fmt_time(p.jam_pinjam),
        _fmt_date(p.tanggal_kembali),
        _fmt_time(p.jam_kembali),
        p.keperluan,
        p.status,
    ]


@admin_required
def laporan_view(request):
    peminjaman, tgl_awal, tgl_akhir, status, ruangan, dosen, prodi = \
        _get_filtered_data(request)

    total = peminjaman.count()
    dipinjam = peminjaman.filter(status='Dipinjam').count()

    context = {
        'data': peminjaman,
        'total': total,
        'dipinjam': dipinjam,
        'dikembalikan': total - dipinjam,
        'tgl_awal': tgl_awal,
        'tgl_akhir': tgl_akhir,
        'status_filter': status,
        'ruangan_filter': ruangan,
        'dosen_filter': dosen,
        'prodi_filter': prodi,
        'lab_list': Laboratorium.objects.all(),
        'dosen_list': Dosen.objects.all(),
        'prodi_list': _prodi_list(),
    }
    return render(request, 'report/laporan.html', context)


@admin_required
def export_excel(request):
    peminjaman, _, _, _, _, _, _ = _get_filtered_data(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Laporan Peminjaman'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col, header in enumerate(EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, p in enumerate(peminjaman, 1):
        for col, value in enumerate(_peminjaman_to_row(i, p), 1):
            cell = ws.cell(row=i + 1, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    for letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="laporan_peminjaman_{timezone.now().date()}.xlsx"'
    wb.save(response)
    return response


@admin_required
def export_csv(request):
    peminjaman, _, _, _, _, _, _ = _get_filtered_data(request)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="laporan_peminjaman_{timezone.now().date()}.csv"'

    writer = csv.writer(response)
    writer.writerow(EXPORT_HEADERS)

    for i, p in enumerate(peminjaman, 1):
        writer.writerow(_peminjaman_to_row(i, p))

    return response
