"""Import data Mahasiswa/Dosen dari file Excel (.xlsx)."""
import io

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color='0D6EFD', end_color='0D6EFD', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
CENTER = Alignment(horizontal='center', vertical='center')


def _read_rows(file, expected_columns):
    """Baca file xlsx; return (rows, error). rows = list dict per baris."""
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
    except Exception:
        return None, 'File tidak valid atau bukan file Excel (.xlsx).'
    ws = wb.active
    if ws.max_row < 2:
        return [], None

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = {}
        empty = True
        for idx, col in enumerate(expected_columns):
            val = row[idx] if idx < len(row) else None
            val = str(val).strip() if val is not None else ''
            record[col] = val
            if val:
                empty = False
        if not empty:
            record['_row'] = i
            rows.append(record)
    return rows, None


def _summary(dibuat, diupdate, dilewati, errors):
    parts = [f'{dibuat} dibuat', f'{diupdate} diupdate', f'{dilewati} dilewati']
    if errors:
        parts.append(f'{len(errors)} error')
    return ', '.join(parts)


def impor_mahasiswa(file):
    """Import mahasiswa; return dict(dibuat, diupdate, dilewati, errors)."""
    from .models import Mahasiswa

    rows, err = _read_rows(file, ['nim', 'nama', 'program_studi'])
    if err:
        return {'error': err}
    dibuat = diupdate = dilewati = 0
    errors = []
    for r in rows:
        if not r['nim'] or not r['nama']:
            dilewati += 1
            errors.append(f"Baris {r['_row']}: NIM/Nama kosong.")
            continue
        data = {
            'nama': r['nama'],
            'program_studi': r['program_studi'] or '-',
        }
        obj, created = Mahasiswa.objects.update_or_create(
            nim=r['nim'], defaults=data
        )
        dibuat += 1 if created else 0
        diupdate += 0 if created else 1
    return {
        'dibuat': dibuat,
        'diupdate': diupdate,
        'dilewati': dilewati,
        'errors': errors,
        'summary': _summary(dibuat, diupdate, dilewati, errors),
    }


def impor_dosen(file):
    """Import dosen; return dict(dibuat, diupdate, dilewati, errors)."""
    from .models import Dosen

    rows, err = _read_rows(file, ['nidn', 'nama'])
    if err:
        return {'error': err}
    dibuat = diupdate = dilewati = 0
    errors = []
    for r in rows:
        if not r['nidn'] or not r['nama']:
            dilewati += 1
            errors.append(f"Baris {r['_row']}: NIDN/Nama kosong.")
            continue
        obj, created = Dosen.objects.update_or_create(
            nidn=r['nidn'], defaults={'nama': r['nama']}
        )
        dibuat += 1 if created else 0
        diupdate += 0 if created else 1
    return {
        'dibuat': dibuat,
        'diupdate': diupdate,
        'dilewati': dilewati,
        'errors': errors,
        'summary': _summary(dibuat, diupdate, dilewati, errors),
    }


def template_mahasiswa():
    """Generate .xlsx kosong berisi header template mahasiswa."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mahasiswa'
    headers = ['NIM', 'Nama', 'Program Studi']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def template_dosen():
    """Generate .xlsx kosong berisi header template dosen."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Dosen'
    headers = ['NIDN', 'Nama']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf